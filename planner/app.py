"""
NC Personal & Business Financial Planning Platform
Main Dash application — entry point, layout, and globally-persistent callbacks.

Architecture:
  - app.layout holds ALL persistent DOM nodes (stores, sidebar, header).
  - Page layouts contain only page-specific DOM.
  - app.py callbacks only reference IDs that are ALWAYS in the layout.
  - Page-specific callbacks live inside each page file using @dash.callback.
"""
import copy
import json
import base64
import io
from pathlib import Path

import dash
from dash import html, dcc, Input, Output, State, ALL, callback_context, no_update
import dash_bootstrap_components as dbc

from planner.config import DATA_DIR, DEFAULT_TAX_YEAR, DEFAULT_STATE
from planner.data_manager import (
    load_project_state, save_project_state, load_tax_rules,
    get_scenarios_list, save_scenario, delete_scenario, duplicate_scenario,
)
from planner.components.sidebar import render_sidebar
from planner.components.header import render_header

# ─────────────────────────────────────────────────────────────────────────────
# App init
# ─────────────────────────────────────────────────────────────────────────────
app = dash.Dash(
    __name__,
    use_pages=True,
    external_stylesheets=[dbc.themes.SLATE],
    suppress_callback_exceptions=True,
)
app.title = "Personal & Business Financial Planner"

# ─────────────────────────────────────────────────────────────────────────────
# Root layout — ALL persistent IDs live here, not in pages
# ─────────────────────────────────────────────────────────────────────────────
app.layout = html.Div(
    [
        dcc.Location(id="url", refresh=False),
        # ── Persistent session stores ──────────────────────────────────────
        dcc.Store(id="project-state-store", storage_type="session"),
        dcc.Store(id="active-scenario-store", data="Baseline", storage_type="session"),
        dcc.Store(id="explain-target-store", data="combined_tax", storage_type="session"),
        # Settings stores consumed by page callbacks
        dcc.Store(id="forecast-horizon-store", data=8, storage_type="session"),
        dcc.Store(id="sensitivity-method-store", data="EBITDA Multiple", storage_type="session"),
        dcc.Store(id="sensitivity-range-store", data=0.20, storage_type="session"),

        # ── Chrome (always visible) ────────────────────────────────────────
        render_sidebar(),
        html.Div(
            [
                render_header(),          # header-scenario-dropdown, save-status-indicator
                dash.page_container,      # page content swapped here on navigation
            ],
            className="main-content",
        ),
    ]
)


# ─────────────────────────────────────────────────────────────────────────────
# Callback: Update page title dynamically
# ─────────────────────────────────────────────────────────────────────────────
@app.callback(
    Output("page-title", "children"),
    Input("url", "pathname")
)
def update_page_title(pathname):
    titles = {
        "/": "Financial Overview",
        "/personal": "Personal Finances",
        "/business": "Business Planning",
        "/taxes": "Tax Planning & Auditing",
        "/networth": "Net Worth Tracking",
        "/valuation": "Business Valuation",
        "/forecast": "Financial Projections & Spreadsheet",
        "/scenarios": "Scenario Manager",
        "/settings": "Settings & Backup"
    }
    return titles.get(pathname, "Financial Planner")



# ─────────────────────────────────────────────────────────────────────────────
# Callback: Synchronize scenario state and dropdown
#           Combines inputs and outputs for active-scenario-store and header-scenario-dropdown
#           to prevent any static dependency cycles.
# ─────────────────────────────────────────────────────────────────────────────
@app.callback(
    Output("project-state-store", "data"),
    Output("header-scenario-dropdown", "options"),
    Output("header-scenario-dropdown", "value"),
    Output("active-scenario-store", "data"),
    Input("header-scenario-dropdown", "value"),
    Input("active-scenario-store", "data"),
    prevent_initial_call=False,
)
def sync_scenario_and_state(dropdown_val, active_scenario):
    ctx = callback_context
    triggered_id = ctx.triggered[0]["prop_id"] if ctx.triggered else ""

    # Defaults
    active_scenario = active_scenario or "Baseline"
    dropdown_val = dropdown_val or "Baseline"

    current_scenario = active_scenario
    if "header-scenario-dropdown" in triggered_id:
        current_scenario = dropdown_val
    elif "active-scenario-store" in triggered_id:
        current_scenario = active_scenario

    scenario_list = get_scenarios_list()
    options = [{"label": s, "value": s} for s in scenario_list]

    state = load_project_state(current_scenario)

    return state, options, current_scenario, current_scenario



# ─────────────────────────────────────────────────────────────────────────────
# Callback: Persist edits from profile/business/valuation inputs & tables
# ─────────────────────────────────────────────────────────────────────────────
@app.callback(
    Output("project-state-store", "data", allow_duplicate=True),
    Output("save-status-indicator", "children"),
    Input({"type": "profile-input",    "field": ALL}, "value"),
    Input({"type": "business-input",   "field": ALL}, "value"),
    Input({"type": "valuation-input",  "field": ALL}, "value"),
    Input("income-table",    "data"),
    Input("expenses-table",  "data"),
    Input("assets-table",    "data"),
    Input("liabilities-table","data"),
    Input("forecast-spreadsheet", "data"),
    State({"type": "profile-input",   "field": ALL}, "id"),
    State({"type": "business-input",  "field": ALL}, "id"),
    State({"type": "valuation-input", "field": ALL}, "id"),
    State("project-state-store", "data"),
    State("active-scenario-store", "data"),
    prevent_initial_call=True,
)
def persist_edits(
    profile_vals, biz_vals, val_vals,
    inc_data, exp_data, ast_data, liab_data, forecast_data,
    profile_ids, biz_ids, val_ids,
    current_state, active_scenario,
):
    if current_state is None:
        return no_update, no_update

    ctx = callback_context
    triggered = ctx.triggered[0]["prop_id"] if ctx.triggered else ""
    if not triggered:
        return no_update, no_update

    new_state = copy.deepcopy(current_state)

    if "profile-input" in triggered:
        for pid, val in zip(profile_ids, profile_vals):
            if val is not None:
                new_state["profile"][pid["field"]] = val

    elif "business-input" in triggered:
        for bid, val in zip(biz_ids, biz_vals):
            if val is not None:
                new_state["business"][bid["field"]] = val

    elif "valuation-input" in triggered:
        mult_map = {
            "revenue_mult": "revenue", "ebitda_mult": "ebitda",
            "net_income_mult": "net_income", "sde_mult": "sde", "fcf_mult": "fcf",
        }
        for vid, val in zip(val_ids, val_vals):
            if val is None:
                continue
            field = vid["field"]
            if field in mult_map:
                new_state["assumptions"]["valuation_multiples"][mult_map[field]] = val
            elif field == "custom_name":
                new_state["assumptions"]["custom_valuation_name"] = val
            elif field == "custom_mult":
                new_state["assumptions"]["custom_valuation_multiplier"] = val

    elif "income-table" in triggered and inc_data is not None:
        new_state["income"] = inc_data
    elif "expenses-table" in triggered and exp_data is not None:
        new_state["expenses"] = exp_data
    elif "assets-table" in triggered and ast_data is not None:
        new_state["assets"] = ast_data
    elif "liabilities-table" in triggered and liab_data is not None:
        new_state["liabilities"] = liab_data
    elif "forecast-spreadsheet" in triggered and forecast_data is not None:
        hist_rows = [r for r in forecast_data if "2025" in str(r.get("Quarter", ""))]
        new_state["forecast"] = hist_rows
        overrides = {}
        for row in forecast_data:
            q = str(row.get("Quarter", ""))
            if "2025" not in q:
                overrides[q] = {k: float(row.get(k, 0) or 0)
                                 for k in ["Revenue", "COGS", "Payroll", "Expenses",
                                           "Capital expenditures", "Owner salary", "Distributions"]}
        new_state["assumptions"]["forecast_overrides"] = overrides

    try:
        save_project_state(new_state, active_scenario)
        label = "● Saved"
    except Exception as e:
        label = f"⚠ {e}"

    return new_state, label


# ─────────────────────────────────────────────────────────────────────────────
# Callbacks: Add rows to editable tables
# ─────────────────────────────────────────────────────────────────────────────
@app.callback(
    Output("income-table", "data", allow_duplicate=True),
    Input("income-table-add-btn", "n_clicks"),
    State("income-table", "data"),
    prevent_initial_call=True,
)
def add_income_row(n, rows):
    rows = rows or []
    if n:
        rows.append({"id": f"inc_{len(rows)+1}", "category": "Other", "description": "New Source", "amount": 0.0})
    return rows


@app.callback(
    Output("expenses-table", "data", allow_duplicate=True),
    Input("expenses-table-add-btn", "n_clicks"),
    State("expenses-table", "data"),
    prevent_initial_call=True,
)
def add_expense_row(n, rows):
    rows = rows or []
    if n:
        rows.append({"id": f"exp_{len(rows)+1}", "category": "Other", "description": "New Expense", "amount": 0.0})
    return rows


@app.callback(
    Output("assets-table", "data", allow_duplicate=True),
    Input("assets-table-add-btn", "n_clicks"),
    State("assets-table", "data"),
    prevent_initial_call=True,
)
def add_asset_row(n, rows):
    rows = rows or []
    if n:
        rows.append({"id": f"ast_{len(rows)+1}", "category": "Other", "description": "New Asset", "value": 0.0, "growth_rate": 0.05})
    return rows


@app.callback(
    Output("liabilities-table", "data", allow_duplicate=True),
    Input("liabilities-table-add-btn", "n_clicks"),
    State("liabilities-table", "data"),
    prevent_initial_call=True,
)
def add_liability_row(n, rows):
    rows = rows or []
    if n:
        rows.append({"id": f"liab_{len(rows)+1}", "category": "Other", "description": "New Liability",
                     "value": 0.0, "interest_rate": 0.05, "monthly_payment": 0.0})
    return rows


# ─────────────────────────────────────────────────────────────────────────────
# Callback: Forecast horizon dropdown → store
# ─────────────────────────────────────────────────────────────────────────────
@app.callback(
    Output("forecast-horizon-store", "data"),
    Input("forecast-horizon-dropdown", "value"),
    prevent_initial_call=True,
)
def update_horizon_store(val):
    return val or 8


# ─────────────────────────────────────────────────────────────────────────────
# Callback: Valuation sensitivity inputs → stores
# ─────────────────────────────────────────────────────────────────────────────
@app.callback(
    Output("sensitivity-method-store", "data"),
    Input("valuation-sensitivity-method", "value"),
    prevent_initial_call=True,
)
def update_sens_method(val):
    return val or "EBITDA Multiple"


@app.callback(
    Output("sensitivity-range-store", "data"),
    Input("valuation-sensitivity-range", "value"),
    prevent_initial_call=True,
)
def update_sens_range(val):
    return val or 0.20


# ─────────────────────────────────────────────────────────────────────────────
# Callback: Export (Settings page)
# ─────────────────────────────────────────────────────────────────────────────
@app.callback(
    Output("settings-download-component", "data"),
    Input("settings-export-json-btn", "n_clicks"),
    Input("settings-export-excel-btn", "n_clicks"),
    State("project-state-store", "data"),
    State("active-scenario-store", "data"),
    prevent_initial_call=True,
)
def export_data(json_c, excel_c, state, active_scenario):
    ctx = callback_context
    triggered = ctx.triggered[0]["prop_id"] if ctx.triggered else ""

    if "json" in triggered and state:
        return dcc.send_string(
            json.dumps(state, indent=2),
            filename=f"finance_planner_{active_scenario}.json",
        )

    if "excel" in triggered and state:
        import openpyxl
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        def add_sheet(title, data):
            if not data:
                return
            ws = wb.create_sheet(title)
            ws.append(list(data[0].keys()))
            for item in data:
                ws.append(list(item.values()))

        ws = wb.create_sheet("Profile")
        for k, v in state["profile"].items():
            ws.append([k, v])

        add_sheet("Income", state["income"])
        add_sheet("Expenses", state["expenses"])
        add_sheet("Assets", state["assets"])
        add_sheet("Liabilities", state["liabilities"])
        add_sheet("Forecast History", state["forecast"])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return dcc.send_bytes(buf.read(), filename=f"financial_model_{active_scenario}.xlsx")

    return no_update


# ─────────────────────────────────────────────────────────────────────────────
# Callback: Import (Settings page)
# ─────────────────────────────────────────────────────────────────────────────
@app.callback(
    Output("settings-import-status", "children"),
    Output("project-state-store", "data", allow_duplicate=True),
    Input("settings-import-upload", "contents"),
    State("settings-import-upload", "filename"),
    State("active-scenario-store", "data"),
    prevent_initial_call=True,
)
def import_data(contents, filename, active_scenario):
    if contents is None:
        return no_update, no_update
    _, content_string = contents.split(",")
    decoded = base64.b64decode(content_string)
    try:
        if not (filename or "").endswith(".json"):
            return html.Div("Only .json files are supported.", style={"color": "#ef4444"}), no_update
        imported = json.loads(decoded.decode("utf-8"))
        for req in ["profile", "business", "assets", "liabilities", "income", "expenses"]:
            if req not in imported:
                return html.Div(f"Missing section: '{req}'", style={"color": "#ef4444"}), no_update
        save_project_state(imported, active_scenario)
        return html.Div("✓ Import successful!", style={"color": "#10b981"}), imported
    except Exception as e:
        return html.Div(f"Error: {e}", style={"color": "#ef4444"}), no_update


# ─────────────────────────────────────────────────────────────────────────────
# Entry point
# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    app.run(debug=True, port=8050)
