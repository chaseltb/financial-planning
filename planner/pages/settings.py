import base64
import io
import json
from datetime import datetime

import dash
from dash import html, dcc, callback, callback_context, Input, Output, State, no_update
import dash_bootstrap_components as dbc

from planner.data_manager import save_project_state

dash.register_page(__name__, path="/settings", title="Settings & Backup")

def layout():
    return dbc.Container(
        [
            dbc.Row(
                [
                    # Rule Settings
                    dbc.Col(
                        html.Div(
                            [
                                html.H4(
                                    [html.I(className="bi bi-sliders me-2 text-info"), "Tax Rules & Display Defaults"],
                                    className="mb-4"
                                ),
                                html.Div(
                                    [
                                        html.Label(
                                            [html.I(className="bi bi-calendar3 me-1 text-muted"), "Filing / Planning Tax Year"],
                                            id="label-tax-year", className="form-label",
                                        ),
                                        dbc.Tooltip("The tax year whose rules and brackets will be used for all calculations.", target="label-tax-year"),
                                        dcc.Dropdown(
                                            id="settings-tax-year",
                                            options=[{"label": "2026 Rules", "value": 2026}],
                                            value=2026,
                                            clearable=False,
                                            className="mb-3",
                                            style={"color": "#0f172a"}
                                        ),

                                        html.Label(
                                            [html.I(className="bi bi-geo-alt me-1 text-muted"), "Planning State"],
                                            id="label-state", className="form-label",
                                        ),
                                        dbc.Tooltip("State whose income tax rules will be applied. Currently supports North Carolina.", target="label-state"),
                                        dcc.Dropdown(
                                            id="settings-state",
                                            options=[{"label": "North Carolina (NC)", "value": "NC"}],
                                            value="NC",
                                            clearable=False,
                                            className="mb-3",
                                            style={"color": "#0f172a"}
                                        ),

                                        html.Label(
                                            [html.I(className="bi bi-palette me-1 text-muted"), "Theme Selection"],
                                            id="label-theme", className="form-label",
                                        ),
                                        dbc.Tooltip("Switch between dark and light interface themes.", target="label-theme"),
                                        dcc.Dropdown(
                                            id="settings-theme",
                                            options=[
                                                {"label": "Dark Mode (Slate)", "value": "dark"},
                                                {"label": "Light Mode (Flatly)", "value": "light"}
                                            ],
                                            value="dark",
                                            clearable=False,
                                            className="mb-3",
                                            style={"color": "#0f172a"}
                                        ),

                                        dbc.Checklist(
                                            options=[
                                                {"label": "Enable Autosave", "value": "autosave"}
                                            ],
                                            value=["autosave"],
                                            id="settings-autosave-toggle",
                                            switch=True,
                                            className="mb-2"
                                        ),
                                        html.Div(
                                            "When off, your edits still apply during this session but "
                                            "aren't written to disk until you click \"Save Now\".",
                                            className="text-muted mb-3", style={"fontSize": "0.8rem"},
                                        ),
                                        dbc.Button(
                                            [html.I(className="bi bi-save me-1"), "Save Now"],
                                            id="settings-save-now-btn", color="secondary", size="sm",
                                            className="mb-2",
                                        ),
                                        html.Div(id="settings-save-now-status", style={"fontSize": "0.85rem"}),
                                    ]
                                )
                            ],
                            className="glass-card mb-4"
                        ),
                        lg=4
                    ),

                    # Backups, Imports, Exports
                    dbc.Col(
                        [
                            html.Div(
                                [
                                    html.H4(
                                        [html.I(className="bi bi-cloud-download me-2 text-info"), "Export Model Configuration"],
                                        className="mb-3"
                                    ),
                                    html.P(
                                        "Download your entire profile, business settings, assets, liabilities, and quarterly forecast as a backup JSON file or an Excel workbook.",
                                        className="text-muted",
                                        style={"fontSize": "0.9rem"}
                                    ),
                                    html.Div(
                                        [
                                            dbc.Button(
                                                [html.I(className="bi bi-filetype-json me-1"), "Download Full JSON Backup"],
                                                id="settings-export-json-btn", color="primary", className="me-2 mb-2"
                                            ),
                                            dbc.Button(
                                                [html.I(className="bi bi-file-earmark-spreadsheet me-1"), "Download Excel Sheet"],
                                                id="settings-export-excel-btn", color="success", className="me-2 mb-2"
                                            ),
                                            dcc.Download(id="settings-download-component")
                                        ],
                                        className="d-flex flex-wrap"
                                    )
                                ],
                                className="glass-card mb-4"
                            ),

                            html.Div(
                                [
                                    html.H4(
                                        [html.I(className="bi bi-cloud-upload me-2 text-info"), "Import Backup Data"],
                                        className="mb-3"
                                    ),
                                    html.P(
                                        "Drag and drop a previously exported JSON backup file here to restore your planning workspace.",
                                        className="text-muted",
                                        style={"fontSize": "0.9rem"}
                                    ),
                                    dcc.Upload(
                                        id="settings-import-upload",
                                        children=html.Div(
                                            [
                                                html.I(className="bi bi-upload me-2"),
                                                "Drag and Drop or ",
                                                html.A("Select Files", style={"color": "var(--accent-blue)"})
                                            ]
                                        ),
                                        style={
                                            "width": "100%",
                                            "height": "100px",
                                            "lineHeight": "100px",
                                            "borderWidth": "2px",
                                            "borderStyle": "dashed",
                                            "borderRadius": "12px",
                                            "textAlign": "center",
                                            "borderColor": "var(--border-glass)",
                                            "backgroundColor": "var(--bg-secondary)",
                                            "cursor": "pointer"
                                        },
                                        multiple=False
                                    ),
                                    html.Div(id="settings-import-status", className="mt-3", style={"fontSize": "0.9rem"})
                                ],
                                className="glass-card mb-4"
                            )
                        ],
                        lg=8
                    )
                ]
            )
        ],
        fluid=True
    )



# settings-theme/settings-autosave-toggle only exist in the Settings page's own
# layout, not in the persistent app-wide chrome. A callback whose Outputs mix a
# page-local target with an always-present one (e.g. theme-store) errors on
# every other page — Dash only silently skips a callback when ALL of its
# outputs are absent, not when some resolve and some don't. So each direction
# below keeps its Output(s) uniformly page-local or uniformly global, and reads
# the store via State (State isn't part of Dash's cycle graph) gated on
# Input("url", "pathname") — url and theme-store/autosave-enabled-store are
# both always present, so this never mixes resolvable and unresolvable Outputs.
# ─────────────────────────────────────────────────────────────────────────────
# Callback: Theme selection (reflects the shared theme-store on page load/nav)
# ─────────────────────────────────────────────────────────────────────────────
@callback(
    Output("settings-theme", "value"),
    Input("url", "pathname"),
    State("theme-store", "data"),
    prevent_initial_call=False,
)
def populate_theme_dropdown(_pathname, theme):
    return theme or "dark"


@callback(
    Output("theme-store", "data"),
    Input("settings-theme", "value"),
    prevent_initial_call=True,
)
def sync_theme_store_from_dropdown(theme):
    return theme or "dark"


# ─────────────────────────────────────────────────────────────────────────────
# Callback: Autosave toggle (reflects/updates the shared autosave-enabled-store)
# ─────────────────────────────────────────────────────────────────────────────
@callback(
    Output("settings-autosave-toggle", "value"),
    Input("url", "pathname"),
    State("autosave-enabled-store", "data"),
    prevent_initial_call=False,
)
def populate_autosave_toggle(_pathname, enabled):
    return ["autosave"] if enabled is not False else []


@callback(
    Output("autosave-enabled-store", "data"),
    Input("settings-autosave-toggle", "value"),
    prevent_initial_call=True,
)
def sync_autosave_store_from_toggle(toggle_val):
    return "autosave" in (toggle_val or [])


# ─────────────────────────────────────────────────────────────────────────────
# Callback: Manual "Save Now" (writes the in-memory state to disk immediately —
# needed when autosave is off, but works regardless)
# ─────────────────────────────────────────────────────────────────────────────
@callback(
    Output("settings-save-now-status", "children"),
    Output("save-status-indicator", "children", allow_duplicate=True),
    Input("settings-save-now-btn", "n_clicks"),
    State("project-state-store", "data"),
    State("active-scenario-store", "data"),
    prevent_initial_call=True,
)
def save_now(n_clicks, state, active_scenario):
    if state is None:
        return html.Span("Nothing to save yet.", className="text-muted"), no_update
    try:
        save_project_state(state, active_scenario)
        timestamp = datetime.now().strftime("%I:%M:%S %p")
        status = html.Span(f"✓ Saved to disk at {timestamp}.", style={"color": "#10b981"})
        header_status = html.Span(
            [html.I(className="bi bi-cloud-check-fill me-1"), f"Saved {timestamp}"],
            style={"color": "#10b981"},
        )
    except Exception as e:
        status = html.Span(f"⚠ Save failed: {e}", style={"color": "#ef4444"})
        header_status = html.Span(
            [html.I(className="bi bi-exclamation-triangle-fill me-1"), f"Save failed: {e}"],
            style={"color": "#ef4444"},
        )
    return status, header_status


# ─────────────────────────────────────────────────────────────────────────────
# Callback: Export (this page's export buttons)
# ─────────────────────────────────────────────────────────────────────────────
@callback(
    Output("settings-download-component", "data"),
    Input("settings-export-json-btn", "n_clicks"),
    Input("settings-export-excel-btn", "n_clicks"),
    State("project-state-store", "data"),
    State("active-scenario-store", "data"),
    prevent_initial_call=True,
)
def export_data(json_c, excel_c, state, active_scenario):
    ctx = callback_context
    triggered = ctx.triggered[0]["prop_id"] if ctx.triggered else ""

    if "json" in triggered and state:
        return dcc.send_string(
            json.dumps(state, indent=2),
            filename=f"finance_planner_{active_scenario}.json",
        )

    if "excel" in triggered and state:
        import openpyxl
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        def add_sheet(title, data):
            if not data:
                return
            ws = wb.create_sheet(title)
            ws.append(list(data[0].keys()))
            for item in data:
                ws.append(list(item.values()))

        ws = wb.create_sheet("Profile")
        for k, v in state["profile"].items():
            ws.append([k, v])

        add_sheet("Income", state["income"])
        add_sheet("Expenses", state["expenses"])
        add_sheet("Assets", state["assets"])
        add_sheet("Liabilities", state["liabilities"])
        add_sheet("Forecast History", state["forecast"])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return dcc.send_bytes(buf.read(), filename=f"financial_model_{active_scenario}.xlsx")

    return no_update


# ─────────────────────────────────────────────────────────────────────────────
# Callback: Import (this page's upload control)
# ─────────────────────────────────────────────────────────────────────────────
@callback(
    Output("settings-import-status", "children"),
    Output("project-state-store", "data", allow_duplicate=True),
    Input("settings-import-upload", "contents"),
    State("settings-import-upload", "filename"),
    State("active-scenario-store", "data"),
    prevent_initial_call=True,
)
def import_data(contents, filename, active_scenario):
    if contents is None:
        return no_update, no_update
    _, content_string = contents.split(",")
    decoded = base64.b64decode(content_string)
    try:
        if not (filename or "").endswith(".json"):
            return html.Div("Only .json files are supported.", style={"color": "#ef4444"}), no_update
        imported = json.loads(decoded.decode("utf-8"))
        for req in ["profile", "business", "assets", "liabilities", "income", "expenses"]:
            if req not in imported:
                return html.Div(f"Missing section: '{req}'", style={"color": "#ef4444"}), no_update
        save_project_state(imported, active_scenario)
        return html.Div("✓ Import successful!", style={"color": "#10b981"}), imported
    except Exception as e:
        return html.Div(f"Error: {e}", style={"color": "#ef4444"}), no_update
